from flask import send_file
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from openpyxl import Workbook
from io import BytesIO
from flask import Flask, jsonify, request, send_file
from flask_cors import CORS
from sqlalchemy import create_engine, text
from reportlab.pdfgen import canvas
from openpyxl import Workbook
import os 
from dotenv import load_dotenv
import smtplib
import ssl
import threading
from email.message import EmailMessage
from datetime import date, timedelta
load_dotenv()
app = Flask(__name__)
CORS(app, resources={r"/*": {"origins": "*"}}, supports_credentials=True)

DATABASE_URL = os.getenv("DATABASE_URL")
engine = create_engine(DATABASE_URL)


@app.route("/")
def home():
    return jsonify({
        "project": "Real3st Shooters Academy BI System",
        "version": "1.0",
        "status": "Running"
    })


@app.route("/students", methods=["GET"])
def get_students():

    with engine.connect() as connection:
        result = connection.execute(text("""
            SELECT
                student_id,
                first_name,
                last_name,
                gender,
                email,
                phone
            FROM students
            ORDER BY student_id;
        """))

        students = []

        for row in result:
            students.append({
                "id": row.student_id,
                "first_name": row.first_name,
                "last_name": row.last_name,
                "gender": row.gender,
                "email": row.email,
                "phone": row.phone
            })

    return jsonify(students)


@app.route("/students", methods=["POST"])
def add_student():

    data = request.get_json()

    with engine.connect() as connection:
        connection.execute(
            text("""
                INSERT INTO students(
                    first_name,
                    last_name,
                    gender,
                    date_of_birth,
                    phone,
                    email,
                    address
                )
                VALUES(
                    :first_name,
                    :last_name,
                    :gender,
                    :date_of_birth,
                    :phone,
                    :email,
                    :address
                )
            """),
            data
        )
        connection.commit()

    return jsonify({"message": "Student added successfully!"}), 201


@app.route("/students/<int:student_id>", methods=["PUT"])
def update_student(student_id):

    data = request.get_json()

    with engine.connect() as connection:
        connection.execute(
            text("""
                UPDATE students
                SET
                    first_name = :first_name,
                    last_name = :last_name,
                    gender = :gender,
                    date_of_birth = :date_of_birth,
                    phone = :phone,
                    email = :email,
                    address = :address
                WHERE student_id = :student_id
            """),
            {
                "student_id": student_id,
                "first_name": data["first_name"],
                "last_name": data["last_name"],
                "gender": data.get("gender"),
                "date_of_birth": data.get("date_of_birth"),
                "phone": data["phone"],
                "email": data["email"],
                "address": data["address"]
            }
        )
        connection.commit()

    return jsonify({"message": "Student updated successfully!"})
@app.route("/students/<int:student_id>", methods=["DELETE"])
def delete_student(student_id):

    with engine.connect() as connection:
        connection.execute(
            text("""
                DELETE FROM students
                WHERE student_id = :student_id
            """),
            {"student_id": student_id}
        )
        connection.commit()

    return jsonify({"message": "Student deleted successfully!"})

@app.route("/reports/students/pdf")
def students_pdf():

    filename = "students_report.pdf"

    c = canvas.Canvas(filename)

    c.setFont("Helvetica-Bold", 16)
    c.drawString(50, 800, "Real3st Shooters Academy")
    c.drawString(50, 780, "Students Report")

    y = 740

    with engine.connect() as conn:

        result = conn.execute(text("""
            SELECT student_id,
                   first_name,
                   last_name,
                   email,
                   phone
            FROM students
            ORDER BY student_id
        """))

        for row in result:

            c.drawString(
                50,
                y,
                f"{row.student_id}  {row.first_name} {row.last_name}  {row.email}"
            )

            y -= 20

            if y < 50:
                c.showPage()
                y = 780

    c.save()

    return send_file(filename, as_attachment=True)

@app.route("/reports/students/excel")
def students_excel():

    wb = Workbook()

    ws = wb.active

    ws.title = "Students"

    ws.append([
        "ID",
        "First Name",
        "Last Name",
        "Email",
        "Phone"
    ])

    with engine.connect() as conn:

        result = conn.execute(text("""
            SELECT student_id,
                   first_name,
                   last_name,
                   email,
                   phone
            FROM students
            ORDER BY student_id
        """))

        for row in result:

            ws.append([
                row.student_id,
                row.first_name,
                row.last_name,
                row.email,
                row.phone
            ])

    filename = "students_report.xlsx"

    wb.save(filename)

    return send_file(filename, as_attachment=True)

@app.route("/courses", methods=["GET"])
def get_courses():

    with engine.connect() as connection:
        result = connection.execute(text("""
            SELECT
                course_id,
                course_name,
                category,
                duration_days,
                price
            FROM courses
            ORDER BY course_id;
        """))

        courses = []

        for row in result:
            courses.append({
                "course_id": row.course_id,
                "course_name": row.course_name,
                "category": row.category,
                "duration_hours": row.duration_days,
                "duration_days": row.duration_days,
                "price": float(row.price)
            })

    return jsonify(courses)
@app.route("/courses", methods=["POST"])
def add_course():

    data = request.json

    with engine.connect() as connection:
        connection.execute(text("""
            INSERT INTO courses (course_name, category, duration_days, price)
            VALUES (:name, :category, :duration, :price)
        """), {
            "name": data["course_name"],
            "category": data.get("category"),
            "duration": data["duration_days"],
            "price": data["price"]
        })

        connection.commit()

    return jsonify({"message": "Course added successfully"})
@app.route("/courses/<int:course_id>", methods=["PUT"])
def update_course(course_id):

    data = request.get_json()

    with engine.connect() as connection:
        connection.execute(
            text("""
                UPDATE courses
                SET
                    course_name = :course_name,
                    category = :category,
                    duration_days = :duration_days,
                    price = :price
                WHERE course_id = :course_id
            """),
            {
                "course_id": course_id,
                "course_name": data["course_name"],
                "category": data.get("category"),
                "duration_days": data["duration_days"],
                "price": data["price"]
            }
        )
        connection.commit()

    return jsonify({"message": "Course updated successfully!"})
@app.route("/courses/<int:course_id>", methods=["DELETE"])
def delete_course(course_id):
    try:
        with engine.begin() as conn:
            conn.execute(
                text("DELETE FROM courses WHERE course_id = :course_id"),
                {"course_id": course_id}
            )

        return jsonify({"message": "Course deleted successfully"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
# ==========================
# PAYMENTS ROUTES
# ==========================

@app.route("/payments", methods=["GET"])
def get_payments():
    try:
        with engine.connect() as conn:
            result = conn.execute(text("""
            SELECT
                 p.payment_id,
                 p.student_id,
                 CONCAT(s.first_name, ' ', s.last_name) AS student_name,
                 p.course_id,
                 c.course_name,
                 p.amount,
                 p.payment_method,
                 p.payment_date,
                 p.status
FROM payments p
JOIN students s
    ON p.student_id = s.student_id
LEFT JOIN courses c
    ON p.course_id = c.course_id
ORDER BY p.payment_id DESC"""))

            payments = []

            for row in result.mappings():
                payments.append({
                    "payment_id": row["payment_id"],
                    "student_id": row["student_id"],
                    "student_name": row["student_name"],
                    "course_id": row["course_id"],
                    "course_name": row["course_name"],
                    "amount": float(row["amount"]),
                    "payment_method": row["payment_method"],
                    "payment_date": str(row["payment_date"]),
                    "status": row["status"]
                })

            return jsonify(payments)

    except Exception as e:
        return jsonify({"error": str(e)}), 500
@app.route("/payments", methods=["POST"])
def add_payment():
    try:
        data = request.get_json()

        if not data:
            return jsonify({
                "error": "No payment data received"
            }), 400

        student_id = data.get("student_id")
        course_id = data.get("course_id")
        amount = data.get("amount")
        payment_method = data.get("payment_method")
        payment_date = data.get("payment_date")
        status = data.get("status", "Completed")

        if not student_id:
            return jsonify({
                "error": "Student is required"
            }), 400

        if not course_id:
            return jsonify({
                "error": "Course is required"
            }), 400

        if amount is None:
            return jsonify({
                "error": "Amount is required"
            }), 400

        if not payment_method:
            return jsonify({
                "error": "Payment method is required"
            }), 400

        if not payment_date:
            return jsonify({
                "error": "Payment date is required"
            }), 400

        with engine.begin() as connection:
            result = connection.execute(
                text("""
                    INSERT INTO payments (
                        student_id,
                        course_id,
                        amount,
                        payment_method,
                        payment_date,
                        status
                    )
                    VALUES (
                        :student_id,
                        :course_id,
                        :amount,
                        :payment_method,
                        :payment_date,
                        :status
                    )
                    RETURNING payment_id
                """),
                {
                    "student_id": int(student_id),
                    "course_id": int(course_id),
                    "amount": float(amount),
                    "payment_method": payment_method,
                    "payment_date": payment_date,
                    "status": status
                }
            )

            payment_id = result.scalar()

        return jsonify({
            "message": "Payment added successfully",
            "payment_id": payment_id
        }), 201

    except ValueError:
        return jsonify({
            "error": "Student, course and amount must have valid values"
        }), 400

    except Exception as error:
        print("ADD PAYMENT ERROR:", error)

        return jsonify({
            "error": str(error)
        }), 500
@app.route("/payments/<int:payment_id>", methods=["PUT"])
def update_payment(payment_id):

    data = request.get_json()

    data["payment_id"] = payment_id

    with engine.begin() as conn:

        conn.execute(text("""
            UPDATE payments
            SET

                student_id=:student_id,
                course_id=:course_id,
                amount=:amount,
                payment_method=:payment_method,
                payment_date=:payment_date,
                status=:status

            WHERE payment_id=:payment_id
        """), data)

    return jsonify({"message":"Payment updated successfully"})


@app.route("/payments/<int:payment_id>", methods=["DELETE"])
def delete_payment(payment_id):

    with engine.begin() as conn:

        conn.execute(text("""
            DELETE FROM payments
            WHERE payment_id=:id
        """), {"id":payment_id})

    return jsonify({"message":"Payment deleted successfully"})

@app.route("/memberships", methods=["GET"])
def get_memberships():

    with engine.connect() as connection:

        result = connection.execute(text("""
            SELECT
                membership_id,
                student_id,
                membership_type,
                start_date,
                expiry_date,
                status
            FROM memberships
            ORDER BY membership_id;
        """))

        memberships = []

        for row in result:
            memberships.append({
                "membership_id": row.membership_id,
                "student_id": row.student_id,
                "membership_type": row.membership_type,
                "start_date": str(row.start_date),
                "expiry_date": str(row.expiry_date),
                "status": row.status
            })

    return jsonify(memberships)
@app.route("/memberships", methods=["POST"])
def add_membership():

    data = request.json

    with engine.begin() as conn:
        conn.execute(
            text("""
                INSERT INTO memberships (
                    student_id,
                    membership_type,
                    start_date,
                    expiry_date,
                    status
                )
                VALUES (
                    :student_id,
                    :membership_type,
                    :start_date,
                    :expiry_date,
                    'Active'
                )
            """),
            {
                "student_id": data["student_id"],
                "membership_type": data["membership_type"],
                "start_date": data["start_date"],
                "expiry_date": data["expiry_date"]
            }
        )

    return jsonify({"message": "Membership added successfully"})
@app.route("/memberships/<int:membership_id>", methods=["PUT"])
def update_membership(membership_id):

    data = request.get_json()

    try:
        with engine.begin() as conn:
            conn.execute(
                text("""
                    UPDATE memberships
                    SET
                        student_id = :student_id,
                        membership_type = :membership_type,
                        start_date = :start_date,
                        expiry_date = :expiry_date
                    WHERE membership_id = :membership_id
                """),
                {
                    "membership_id": membership_id,
                    "student_id": data["student_id"],
                    "membership_type": data["membership_type"],
                    "start_date": data["start_date"],
                    "expiry_date": data["expiry_date"]
                }
            )

        return jsonify({"message": "Membership updated successfully"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/memberships/<int:membership_id>", methods=["DELETE"])
def delete_membership(membership_id):
    try:
        with engine.begin() as conn:
            conn.execute(
                text("DELETE FROM memberships WHERE membership_id = :membership_id"),
                {"membership_id": membership_id}
            )

        return jsonify({"message": "Membership deleted successfully"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/memberships/send-reminders", methods=["POST"])
def send_membership_reminders():
    """Send email reminders for memberships expiring within `days` days.

    Expects JSON body: { "days": 7 }
    SMTP configuration is read from environment variables:
      SMTP_HOST, SMTP_PORT, SMTP_USER, SMTP_PASS, FROM_EMAIL
    If SMTP_USER/SMTP_PASS are not provided, will attempt to send without auth (localhost).
    """
    payload = request.get_json() or {}
    days = int(payload.get("days", 7))

    today = date.today()
    target = today + timedelta(days=days)

    today_str = today.isoformat()
    target_str = target.isoformat()

    # query memberships that expire between today and target
    with engine.connect() as conn:
        result = conn.execute(text("""
            SELECT
                m.membership_id,
                m.expiry_date,
                m.membership_type,
                s.student_id,
                s.first_name,
                s.last_name,
                s.email
            FROM memberships m
            JOIN students s ON m.student_id = s.student_id
            WHERE m.expiry_date BETWEEN :today AND :target
            ORDER BY m.expiry_date
        """), {"today": today_str, "target": target_str})

        rows = [dict(r._mapping) for r in result]

    if not rows:
        return jsonify({"message": f"No memberships expiring within {days} days."})

    # SMTP config
    smtp_host = os.environ.get("SMTP_HOST", "localhost")
    smtp_port = int(os.environ.get("SMTP_PORT", 25))
    smtp_user = os.environ.get("SMTP_USER")
    smtp_pass = os.environ.get("SMTP_PASS")
    from_email = os.environ.get("FROM_EMAIL", smtp_user or f"no-reply@{smtp_host}")

    sent = []
    failed = []

    # establish SMTP connection once when possible
    try:
        if smtp_user and smtp_pass:
            context = ssl.create_default_context()
            server = smtplib.SMTP(smtp_host, smtp_port, timeout=20)
            server.starttls(context=context)
            server.login(smtp_user, smtp_pass)
        else:
            server = smtplib.SMTP(smtp_host, smtp_port, timeout=20)
    except Exception as e:
        return jsonify({"error": f"Failed to connect to SMTP server: {e}"}), 500

    for r in rows:
        to_email = r.get("email")
        if not to_email:
            failed.append({"id": r.get("membership_id"), "reason": "no email"})
            continue

        student_name = f"{r.get('first_name') or ''} {r.get('last_name') or ''}".strip()
        subj = f"Membership expiring in {days} days"
        body = f"Dear {student_name or 'Member'},\n\nYour {r.get('membership_type')} membership (ID: {r.get('membership_id')}) is set to expire on {r.get('expiry_date')}. Please renew to continue enjoying benefits.\n\nBest regards,\nReal3st Shooters Academy"

        msg = EmailMessage()
        msg["From"] = from_email
        msg["To"] = to_email
        msg["Subject"] = subj
        msg.set_content(body)

        try:
            server.send_message(msg)
            sent.append({"id": r.get("membership_id"), "email": to_email})
        except Exception as e:
            failed.append({"id": r.get("membership_id"), "email": to_email, "error": str(e)})

    try:
        server.quit()
    except Exception:
        pass

    return jsonify({"sent": sent, "failed": failed})

@app.route("/login", methods=["POST"])
def login():

    data = request.get_json()

    with engine.connect() as conn:
        user = conn.execute(
            text("""
                SELECT
                    username,
                    password,
                    full_name,
                    role
                FROM users
                WHERE username = :username
            """),
            {"username": data["username"]}
        ).fetchone()

    if user is None or user.password != data["password"]:
        return jsonify({
            "success": False,
            "message": "Invalid username or password."
        }), 401

    return jsonify({
        "success": True,
        "message": f"Welcome, {user.full_name}",
        "user": {
            "username": user.username,
            "full_name": user.full_name,
            "role": user.role
        }
    })
@app.route("/users", methods=["GET"])
def get_users():
    with engine.connect() as conn:
        result = conn.execute(text("""
            SELECT
                user_id,
                username,
                full_name,
                role
            FROM users
            ORDER BY user_id
        """))

        users = [dict(row._mapping) for row in result]

    return jsonify(users)


@app.route("/users", methods=["POST"])
def add_user():
    data = request.get_json() or {}

    required_fields = [
        "username",
        "password",
        "full_name",
        "role"
    ]

    for field in required_fields:
        if not data.get(field):
            return jsonify({
                "error": f"{field} is required."
            }), 400

    try:
        with engine.begin() as conn:
            conn.execute(
                text("""
                    INSERT INTO users (
                        username,
                        password,
                        full_name,
                        role
                    )
                    VALUES (
                        :username,
                        :password,
                        :full_name,
                        :role
                    )
                """),
                {
                    "username": data["username"],
                    "password": data["password"],
                    "full_name": data["full_name"],
                    "role": data["role"]
                }
            )

        return jsonify({
            "message": "User added successfully."
        }), 201

    except Exception as error:
        return jsonify({
            "error": str(error)
        }), 500


@app.route("/users/<int:user_id>", methods=["PUT"])
def update_user(user_id):
    data = request.get_json() or {}

    try:
        with engine.begin() as conn:

            if data.get("password"):
                conn.execute(
                    text("""
                        UPDATE users
                        SET
                            username = :username,
                            password = :password,
                            full_name = :full_name,
                            role = :role
                        WHERE user_id = :user_id
                    """),
                    {
                        "user_id": user_id,
                        "username": data["username"],
                        "password": data["password"],
                        "full_name": data["full_name"],
                        "role": data["role"]
                    }
                )

            else:
                conn.execute(
                    text("""
                        UPDATE users
                        SET
                            username = :username,
                            full_name = :full_name,
                            role = :role
                        WHERE user_id = :user_id
                    """),
                    {
                        "user_id": user_id,
                        "username": data["username"],
                        "full_name": data["full_name"],
                        "role": data["role"]
                    }
                )

        return jsonify({
            "message": "User updated successfully."
        })

    except Exception as error:
        return jsonify({
            "error": str(error)
        }), 500


@app.route("/users/<int:user_id>", methods=["DELETE"])
def delete_user(user_id):
    data = request.get_json(silent=True) or {}
    logged_in_username = data.get("logged_in_username")

    with engine.connect() as conn:
        user = conn.execute(
            text("""
                SELECT username
                FROM users
                WHERE user_id = :user_id
            """),
            {"user_id": user_id}
        ).fetchone()

    if user is None:
        return jsonify({
            "error": "User not found."
        }), 404

    if user.username == logged_in_username:
        return jsonify({
            "error": "You cannot delete your own account."
        }), 400

    with engine.begin() as conn:
        conn.execute(
            text("""
                DELETE FROM users
                WHERE user_id = :user_id
            """),
            {"user_id": user_id}
        )

    return jsonify({
        "message": "User deleted successfully."
    })

@app.route("/instructors", methods=["GET"])
def get_instructors():
    with engine.connect() as conn:
        result = conn.execute(text("""
            SELECT
                instructor_id,
                full_name,
                specialty,
                phone,
                email
            FROM instructors
            ORDER BY instructor_id;
        """))

        instructors = [dict(row._mapping) for row in result]
        return jsonify(instructors)
@app.route("/certificates", methods=["GET"])
def get_certificates():
    with engine.connect() as conn:
        result = conn.execute(text("""
            SELECT *
            FROM certificates
            ORDER BY certificate_id;
        """))

        certificates = [dict(row._mapping) for row in result]
        return jsonify(certificates)
@app.route("/certificates", methods=["POST"])
def add_certificate():
    data = request.get_json() or {}

    required_fields = [
        "student_id",
        "course_id",
        "certificate_number",
        "issue_date",
        "expiry_date"
    ]

    for field in required_fields:
        if not data.get(field):
            return jsonify({
                "error": f"{field} is required."
            }), 400

    try:
        with engine.begin() as conn:
            conn.execute(
                text("""
                    INSERT INTO certificates (
                        student_id,
                        course_id,
                        certificate_number,
                        issue_date,
                        expiry_date
                    )
                    VALUES (
                        :student_id,
                        :course_id,
                        :certificate_number,
                        :issue_date,
                        :expiry_date
                    )
                """),
                {
                    "student_id": data["student_id"],
                    "course_id": data["course_id"],
                    "certificate_number": data["certificate_number"],
                    "issue_date": data["issue_date"],
                    "expiry_date": data["expiry_date"]
                }
            )

        return jsonify({
            "message": "Certificate added successfully."
        }), 201

    except Exception as error:
        return jsonify({
            "error": str(error)
        }), 500


@app.route("/certificates/<int:certificate_id>", methods=["PUT"])
def update_certificate(certificate_id):
    data = request.get_json() or {}

    try:
        with engine.begin() as conn:
            result = conn.execute(
                text("""
                    UPDATE certificates
                    SET
                        student_id = :student_id,
                        course_id = :course_id,
                        certificate_number = :certificate_number,
                        issue_date = :issue_date,
                        expiry_date = :expiry_date
                    WHERE certificate_id = :certificate_id
                """),
                {
                    "certificate_id": certificate_id,
                    "student_id": data["student_id"],
                    "course_id": data["course_id"],
                    "certificate_number": data["certificate_number"],
                    "issue_date": data["issue_date"],
                    "expiry_date": data["expiry_date"]
                }
            )

        if result.rowcount == 0:
            return jsonify({
                "error": "Certificate not found."
            }), 404

        return jsonify({
            "message": "Certificate updated successfully."
        })

    except Exception as error:
        return jsonify({
            "error": str(error)
        }), 500


@app.route("/certificates/<int:certificate_id>", methods=["DELETE"])
def delete_certificate(certificate_id):
    try:
        with engine.begin() as conn:
            result = conn.execute(
                text("""
                    DELETE FROM certificates
                    WHERE certificate_id = :certificate_id
                """),
                {
                    "certificate_id": certificate_id
                }
            )

        if result.rowcount == 0:
            return jsonify({
                "error": "Certificate not found."
            }), 404

        return jsonify({
            "message": "Certificate deleted successfully."
        })

    except Exception as error:
        return jsonify({
            "error": str(error)
        }), 500

@app.route("/firearms_qualifications", methods=["GET"])
def get_firearms_qualifications():
    with engine.connect() as connection:
        result = connection.execute(text("""
            SELECT
                qualification_id,
                student_id,
                qualification_name,
                qualification_level,
                issue_date,
                instructor
            FROM firearms_qualifications
            ORDER BY qualification_id
        """))

        qualifications = [
            {
                "qualification_id": row.qualification_id,
                "student_id": row.student_id,
                "qualification_name": row.qualification_name,
                "qualification_level": row.qualification_level or "",
                "issue_date": str(row.issue_date) if row.issue_date else "",
                "instructor": row.instructor or ""
            }
            for row in result
        ]

    return jsonify(qualifications) 


@app.route("/firearms_qualifications", methods=["POST"])
def add_firearms_qualification():
    data = request.get_json() or {}

    required_fields = [
        "student_id",
        "qualification_name",
        "qualification_level",
        "issue_date",
        "instructor"
    ]

    for field in required_fields:
        if not data.get(field):
            return jsonify({
                "error": f"{field} is required."
            }), 400

    try:
        with engine.begin() as conn:
            conn.execute(
                text("""
                    INSERT INTO firearms_qualifications (
                        student_id,
                        qualification_name,
                        qualification_level,
                        issue_date,
                        expiry_date,
                        instructor
                    )
                    VALUES (
                        :student_id,
                        :qualification_name,
                        :qualification_level,
                        :issue_date,
                        :expiry_date,
                        :instructor
                    )
                """),
                {
                    "student_id": data["student_id"],
                    "qualification_name": data["qualification_name"],
                    "qualification_level": data["qualification_level"],
                    "issue_date": data["issue_date"],
                    "expiry_date": data.get("expiry_date"),
                    "instructor": data["instructor"]
                }
            )

        return jsonify({
            "message": "Qualification added successfully."
        }), 201

    except Exception as error:
        return jsonify({
            "error": str(error)
        }), 500


@app.route("/firearms_qualifications/<int:qualification_id>", methods=["PUT"])
def update_firearms_qualification(qualification_id):
    data = request.get_json() or {}

    try:
        with engine.begin() as conn:
            result = conn.execute(
                text("""
                    UPDATE firearms_qualifications
                    SET
                        student_id = :student_id,
                        qualification_name = :qualification_name,
                        qualification_level = :qualification_level,
                        issue_date = :issue_date,
                        expiry_date = :expiry_date,
                        instructor = :instructor
                    WHERE qualification_id = :qualification_id
                """),
                {
                    "qualification_id": qualification_id,
                    "student_id": data.get("student_id"),
                    "qualification_name": data.get("qualification_name"),
                    "qualification_level": data.get("qualification_level"),
                    "issue_date": data.get("issue_date"),
                    "expiry_date": data.get("expiry_date"),
                    "instructor": data.get("instructor")
                }
            )

        if result.rowcount == 0:
            return jsonify({
                "error": "Qualification not found."
            }), 404

        return jsonify({
            "message": "Qualification updated successfully."
        })

    except Exception as error:
        return jsonify({
            "error": str(error)
        }), 500


@app.route("/firearms_qualifications/<int:qualification_id>", methods=["DELETE"])
def delete_firearms_qualification(qualification_id):
    try:
        with engine.begin() as conn:
            result = conn.execute(
                text("""
                    DELETE FROM firearms_qualifications
                    WHERE qualification_id = :qualification_id
                """),
                {"qualification_id": qualification_id}
            )

        if result.rowcount == 0:
            return jsonify({"error": "Qualification not found."}), 404

        return jsonify({"message": "Qualification deleted successfully."})

    except Exception as error:
        return jsonify({"error": str(error)}), 500


def schedule_membership_reminders(interval_hours: int = 24, days: int = 7):
    def run_reminder():
        with app.app_context():
            try:
                with app.test_request_context(json={"days": days}):
                    send_membership_reminders()
            except Exception as e:
                print(f"[scheduler] Failed to send membership reminders: {e}")
        threading.Timer(interval_hours * 3600, run_reminder).start()

if __name__ == "__main__":
    reminder_interval = int(os.environ.get("REMINDER_INTERVAL_HOURS", 24))
    reminder_days = int(os.environ.get("REMINDER_DAYS", 7))
    schedule_membership_reminders(reminder_interval, reminder_days)
    app.run(debug=True)
@app.route("/reports/students/pdf", methods=["GET"])
def export_students_pdf():
    pdf_buffer = BytesIO()

    pdf = canvas.Canvas(pdf_buffer, pagesize=A4)
    page_width, page_height = A4

    pdf.setTitle("Students Report")
    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawString(50, page_height - 50, "Real3st Shooters Academy")
    pdf.setFont("Helvetica-Bold", 13)
    pdf.drawString(50, page_height - 75, "Students Report")

    y = page_height - 110

    pdf.setFont("Helvetica-Bold", 9)
    pdf.drawString(40, y, "ID")
    pdf.drawString(70, y, "Full Name")
    pdf.drawString(220, y, "Email")
    pdf.drawString(400, y, "Phone")

    y -= 18

    with engine.connect() as connection:
        result = connection.execute(text("""
            SELECT
                student_id,
                first_name,
                last_name,
                email,
                phone
            FROM students
            ORDER BY student_id
        """))

        pdf.setFont("Helvetica", 8)

        for row in result:
            if y < 50:
                pdf.showPage()
                y = page_height - 50

            full_name = f"{row.first_name} {row.last_name}"

            pdf.drawString(40, y, str(row.student_id))
            pdf.drawString(70, y, full_name[:28])
            pdf.drawString(220, y, str(row.email or "")[:32])
            pdf.drawString(400, y, str(row.phone or "")[:20])

            y -= 16

    pdf.save()
    pdf_buffer.seek(0)

    return send_file(
        pdf_buffer,
        as_attachment=True,
        download_name="students_report.pdf",
        mimetype="application/pdf"
    )
@app.route("/reports/students/excel", methods=["GET"])
def export_students_excel():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Students"

    worksheet.append([
        "Student ID",
        "First Name",
        "Last Name",
        "Gender",
        "Email",
        "Phone"
    ])

    with engine.connect() as connection:
        result = connection.execute(text("""
            SELECT
                student_id,
                first_name,
                last_name,
                gender,
                email,
                phone
            FROM students
            ORDER BY student_id
        """))

        for row in result:
            worksheet.append([
                row.student_id,
                row.first_name,
                row.last_name,
                row.gender,
                row.email,
                row.phone
            ])

    excel_buffer = BytesIO()
    workbook.save(excel_buffer)
    excel_buffer.seek(0)

    return send_file(
        excel_buffer,
        as_attachment=True,
        download_name="students_report.xlsx",
        mimetype=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )
@app.route("/reports/payments/pdf", methods=["GET"])
def export_payments_pdf():
    pdf_buffer = BytesIO()
    pdf = canvas.Canvas(pdf_buffer, pagesize=A4)

    width, height = A4

    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawString(50, height - 50, "Real3st Shooters Academy")

    pdf.setFont("Helvetica-Bold", 13)
    pdf.drawString(50, height - 75, "Payments Report")

    y = height - 110

    pdf.setFont("Helvetica-Bold", 9)
    pdf.drawString(35, y, "ID")
    pdf.drawString(65, y, "Student")
    pdf.drawString(190, y, "Amount")
    pdf.drawString(260, y, "Date")
    pdf.drawString(350, y, "Method")

    y -= 18

    with engine.connect() as connection:
        result = connection.execute(text("""
            SELECT
                p.payment_id,
                s.first_name,
                s.last_name,
                p.amount,
                p.payment_date,
                p.payment_method
            FROM payments p
            JOIN students s
                ON p.student_id = s.student_id
            ORDER BY p.payment_id
        """))

        pdf.setFont("Helvetica", 8)

        for row in result:
            if y < 50:
                pdf.showPage()
                y = height - 50

            student_name = f"{row.first_name} {row.last_name}"

            pdf.drawString(35, y, str(row.payment_id))
            pdf.drawString(65, y, student_name[:24])
            pdf.drawString(190, y, str(row.amount))
            pdf.drawString(260, y, str(row.payment_date))
            pdf.drawString(350, y, str(row.payment_method))

            y -= 16

    pdf.save()
    pdf_buffer.seek(0)

    return send_file(
        pdf_buffer,
        as_attachment=True,
        download_name="payments_report.pdf",
        mimetype="application/pdf"
    )


@app.route("/reports/payments/excel", methods=["GET"])
def export_payments_excel():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Payments"

    worksheet.append([
        "Payment ID",
        "Student",
        "Amount",
        "Payment Date",
        "Payment Method"
    ])

    with engine.connect() as connection:
        result = connection.execute(text("""
            SELECT
                p.payment_id,
                s.first_name,
                s.last_name,
                p.amount,
                p.payment_date,
                p.payment_method
            FROM payments p
            JOIN students s
                ON p.student_id = s.student_id
            ORDER BY p.payment_id
        """))

        for row in result:
            worksheet.append([
                row.payment_id,
                f"{row.first_name} {row.last_name}",
                float(row.amount),
                str(row.payment_date),
                row.payment_method
            ])

    excel_buffer = BytesIO()
    workbook.save(excel_buffer)
    excel_buffer.seek(0)

    return send_file(
        excel_buffer,
        as_attachment=True,
        download_name="payments_report.xlsx",
        mimetype=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )


@app.route("/reports/memberships/pdf", methods=["GET"])
def export_memberships_pdf():
    pdf_buffer = BytesIO()
    pdf = canvas.Canvas(pdf_buffer, pagesize=A4)

    width, height = A4

    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawString(50, height - 50, "Real3st Shooters Academy")

    pdf.setFont("Helvetica-Bold", 13)
    pdf.drawString(50, height - 75, "Memberships Report")

    y = height - 110

    pdf.setFont("Helvetica-Bold", 9)
    pdf.drawString(35, y, "ID")
    pdf.drawString(65, y, "Student ID")
    pdf.drawString(130, y, "Type")
    pdf.drawString(240, y, "Start Date")
    pdf.drawString(330, y, "Expiry Date")
    pdf.drawString(430, y, "Status")

    y -= 18

    with engine.connect() as connection:
        result = connection.execute(text("""
            SELECT
                membership_id,
                student_id,
                membership_type,
                start_date,
                expiry_date,
                status
            FROM memberships
            ORDER BY membership_id
        """))

        pdf.setFont("Helvetica", 8)

        for row in result:
            if y < 50:
                pdf.showPage()
                y = height - 50

            pdf.drawString(35, y, str(row.membership_id))
            pdf.drawString(65, y, str(row.student_id))
            pdf.drawString(130, y, str(row.membership_type))
            pdf.drawString(240, y, str(row.start_date))
            pdf.drawString(330, y, str(row.expiry_date))
            pdf.drawString(430, y, str(row.status))

            y -= 16

    pdf.save()
    pdf_buffer.seek(0)

    return send_file(
        pdf_buffer,
        as_attachment=True,
        download_name="memberships_report.pdf",
        mimetype="application/pdf"
    )


@app.route("/reports/memberships/excel", methods=["GET"])
def export_memberships_excel():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Memberships"

    worksheet.append([
        "Membership ID",
        "Student ID",
        "Membership Type",
        "Start Date",
        "Expiry Date",
        "Status"
    ])

    with engine.connect() as connection:
        result = connection.execute(text("""
            SELECT
                membership_id,
                student_id,
                membership_type,
                start_date,
                expiry_date,
                status
            FROM memberships
            ORDER BY membership_id
        """))

        for row in result:
            worksheet.append([
                row.membership_id,
                row.student_id,
                row.membership_type,
                str(row.start_date),
                str(row.expiry_date),
                row.status
            ])

    excel_buffer = BytesIO()
    workbook.save(excel_buffer)
    excel_buffer.seek(0)

    return send_file(
        excel_buffer,
        as_attachment=True,
        download_name="memberships_report.xlsx",
        mimetype=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )
@app.route("/dashboard/revenue-chart", methods=["GET"])
def revenue_chart():
    try:
        with engine.connect() as conn:
            result = conn.execute(text("""
                SELECT
                    TO_CHAR(payment_date, 'Mon') AS month,
                    EXTRACT(MONTH FROM payment_date) AS month_number,
                    SUM(amount) AS revenue
                FROM payments
                WHERE status = 'Completed'
                GROUP BY
                    EXTRACT(MONTH FROM payment_date),
                    TO_CHAR(payment_date, 'Mon')
                ORDER BY month_number
            """))

            data = []

            for row in result.mappings():
                data.append({
                    "month": row["month"],
                    "revenue": float(row["revenue"])
                })

            return jsonify(data)

    except Exception as error:
        return jsonify({
            "error": str(error)
        }), 500


@app.route("/dashboard/popular-course", methods=["GET"])
def popular_course():
    try:
        with engine.connect() as conn:
            result = conn.execute(text("""
                SELECT
                    c.course_name,
                    COUNT(p.payment_id) AS total
                FROM payments p
                JOIN courses c
                    ON p.course_id = c.course_id
                WHERE p.status = 'Completed'
                GROUP BY c.course_name
                ORDER BY total DESC
                LIMIT 1
            """))

            row = result.mappings().first()

            if row:
                return jsonify({
                    "course": row["course_name"],
                    "students": row["total"]
                })

            return jsonify({
                "course": "None",
                "students": 0
            })

    except Exception as error:
        return jsonify({
            "error": str(error)
        }), 500
@app.route("/dashboard", methods=["GET"])
def dashboard_summary():
    try:
        with engine.connect() as conn:

            total_students = conn.execute(
                text("SELECT COUNT(*) FROM students")
            ).scalar() or 0

            total_courses = conn.execute(
                text("SELECT COUNT(*) FROM courses")
            ).scalar() or 0

            total_memberships = conn.execute(
                text("SELECT COUNT(*) FROM memberships")
            ).scalar() or 0

            total_payments = conn.execute(
                text("SELECT COUNT(*) FROM payments")
            ).scalar() or 0

            total_revenue = conn.execute(
                text("""
                    SELECT COALESCE(SUM(amount), 0)
                    FROM payments
                    WHERE LOWER(status) = 'completed'
                """)
            ).scalar() or 0

        return jsonify({
            "total_students": total_students,
            "total_courses": total_courses,
            "total_memberships": total_memberships,
            "total_payments": total_payments,
            "total_revenue": float(total_revenue)
        })

    except Exception as error:
        print("DASHBOARD ERROR:", error)

        return jsonify({
            "error": str(error)
        }), 500
@app.route("/dashboard/revenue-by-course", methods=["GET"])
def revenue_by_course():
    try:
        with engine.connect() as conn:
            result = conn.execute(text("""
                SELECT
                    c.course_name,
                    COALESCE(SUM(p.amount), 0) AS revenue
                FROM courses c
                LEFT JOIN payments p
                    ON c.course_id = p.course_id
                    AND LOWER(p.status) = 'completed'
                GROUP BY c.course_id, c.course_name
                ORDER BY revenue DESC
            """))

            data = []

            for row in result.mappings():
                data.append({
                    "course": row["course_name"],
                    "revenue": float(row["revenue"])
                })

            return jsonify(data)

    except Exception as error:
        return jsonify({
            "error": str(error)
        }), 500
@app.route("/dashboard/instructor-workload", methods=["GET"])
def instructor_workload():
    try:
        with engine.connect() as conn:
            result = conn.execute(text("""
                SELECT
                    i.instructor_id,
                    i.full_name AS instructor_name,
                    COUNT(fq.qualification_id) AS total_qualifications
                FROM instructors i
                LEFT JOIN firearms_qualifications fq
                    ON LOWER(TRIM(i.full_name)) =
                       LOWER(TRIM(fq.instructor))
                GROUP BY
                    i.instructor_id,
                    i.full_name
                ORDER BY
                    total_qualifications DESC,
                    i.full_name ASC
            """))

            data = []

            for row in result.mappings():
                data.append({
                    "instructor_id": row["instructor_id"],
                    "instructor_name": row["instructor_name"],
                    "workload": int(row["total_qualifications"])
                })

            return jsonify(data)

    except Exception as error:
        print("INSTRUCTOR WORKLOAD ERROR:", error)

        return jsonify({
            "error": str(error)
        }), 500
@app.route("/dashboard/membership-renewals", methods=["GET"])
def membership_renewals():
    try:
        with engine.connect() as conn:
            result = conn.execute(text("""
                SELECT
                    m.membership_id,
                    m.student_id,
                    m.membership_type,
                    m.end_date,
                    m.status
                FROM memberships m
                WHERE m.end_date IS NOT NULL
                  AND m.end_date <= CURRENT_DATE + INTERVAL '30 days'
                ORDER BY m.end_date ASC
            """))

            data = []

            for row in result.mappings():
                data.append({
                    "membership_id": row["membership_id"],
                    "student_id": row["student_id"],
                    "membership_type": row["membership_type"],
                    "end_date": str(row["end_date"]),
                    "status": row["status"]
                })

            return jsonify(data)

    except Exception as error:
        return jsonify({
            "error": str(error)
        }), 500
@app.route("/dashboard/division-revenue", methods=["GET"])
def division_revenue():
    try:
        with engine.connect() as conn:
            result = conn.execute(text("""
                SELECT
                    c.category,
                    COALESCE(SUM(p.amount), 0) AS revenue
                FROM courses c
                LEFT JOIN payments p
                    ON c.course_id = p.course_id
                    AND LOWER(p.status) = 'completed'
                GROUP BY c.category
                ORDER BY revenue DESC
            """))

            data = []

            for row in result.mappings():
                data.append({
                    "division": row["category"],
                    "revenue": float(row["revenue"])
                })

            return jsonify(data)

    except Exception as error:
        print("DIVISION REVENUE ERROR:", error)

        return jsonify({
            "error": str(error)
        }), 500
@app.route("/dashboard/students-trained", methods=["GET"])
def students_trained():
    try:
        with engine.connect() as conn:

            total = conn.execute(text("""
                SELECT COUNT(DISTINCT student_id)
                FROM certificates
            """)).scalar() or 0

        return jsonify({
            "students_trained": int(total)
        })

    except Exception as error:
        print("STUDENTS TRAINED ERROR:", error)

        return jsonify({
            "error": str(error)
        }), 500
if __name__ == "__main__":
    app.run(debug=True)